from datetime import datetime
from dataclasses import asdict
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from models import Expense


class ExpenseTracker:
    def __init__(self):
        self.filename = "expenses.json"
        self.expenses = self.load_data()

    def get_trend_prediction(self):
        y = self.get_daily_totals_sequence()
        n = len(y)

        if n < 3:
            return None

        x = list(range(1, n+1))
        sum_x = sum(x)
        sum_y = sum(y)
        sum_xx = sum(val**2 for val in x)
        sum_xy = sum(x_val * y_val for x_val, y_val in zip(x, y))

        denominator = n * sum_xx - (sum_x ** 2)
        if denominator == 0:
            k = 0
        else:
            k = (n * sum_xy - sum_x * sum_y) / denominator
        b = (sum_y - k * sum_x) / n
        next_day = n + 1
        predicted_tomorrow = k * next_day + b
        predicted_month_total = sum(max(0, k * day + b) for day in range(next_day, next_day + 30))

        return {
            "trend_direction": "растет 📈" if k > 0 else ("падает 📉" if k < 0 else "стабилен ➡️"),
            "slope_k": round(k, 2),
            "predicted_tomorrow": round(max(0, predicted_tomorrow), 2),
            "predicted_month_total": round(predicted_month_total, 2)
        }

    def get_daily_totals_sequence(self):
        if not self.expenses:
            return []
        daily_map = {}
        for item in self.expenses:
            day_str = item.date[:10]
            if day_str not in daily_map:
                daily_map[day_str] = 0
            daily_map[day_str] += item.amount

        sorted_days = sorted(daily_map.keys())
        sequence = [daily_map[day] for day in sorted_days]
        return sequence

    def get_analytics(self):
        if not self.expenses:
            return {"avg_check": 0,
                    "avg_per_days": 0,
                    "total_days": 0,}
        total_sum = self.get_total()
        avg_check = total_sum / len(self.expenses)
        unique_days = set()
        for item in self.expenses:
            day_str = item.date[:10]
            unique_days.add(day_str)
        total_days = len(unique_days)
        avg_per_day = total_sum / total_days if total_days > 0 else 0
        return {
            "avg_check": round(avg_check, 2),
            "avg_per_days": round(avg_per_day, 2),
            "total_days": total_days
        }

    def export_to_excel(self, export_filename="expenses_report.xlsx"):
        if not self.expenses:
            return False

        wb = openpyxl.Workbook()
        ws_dash = wb.active
        ws_dash.title = "Дашборд"
        ws_data = wb.create_sheet(title="Все расходы")

        ws_dash.views.sheetView[0].showGridLines = True
        ws_data.views.sheetView[0].showGridLines = True

        font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)

        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
        fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        thin_side = Side(border_style="thin", color="D9D9D9")
        border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=Side(border_style="thin", color="1F4E78"),
                              bottom=Side(border_style="double", color="1F4E78"))

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        headers_data = ["№ п/п", "Дата и время", "Категория", "Сумма (руб.)"]
        for col_idx, text in enumerate(headers_data, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        sorted_expenses = sorted(self.expenses, key=lambda x: x.date)

        for row_idx, exp in enumerate(sorted_expenses, 2):
            c1 = ws_data.cell(row=row_idx, column=1, value=row_idx - 1)
            c2 = ws_data.cell(row=row_idx, column=2, value=exp.date)
            c3 = ws_data.cell(row=row_idx, column=3, value=exp.category)
            c4 = ws_data.cell(row=row_idx, column=4, value=exp.amount)

            c1.alignment = align_center
            c2.alignment = align_center
            c3.alignment = align_left
            c4.alignment = align_right
            c4.number_format = '#,##0" руб."'

            for c in [c1, c2, c3, c4]:
                c.font = font_regular
                c.border = border_data
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

        total_row_idx = len(sorted_expenses) + 2
        ws_data.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=3)
        total_label = ws_data.cell(row=total_row_idx, column=1, value="Итого:")
        total_label.font = font_bold
        total_label.alignment = align_right

        total_val = ws_data.cell(row=total_row_idx, column=4, value=f"=SUM(D2:D{total_row_idx - 1})")
        total_val.font = font_bold
        total_val.alignment = align_right
        total_val.number_format = '#,##0" руб."'

        for col_idx in range(1, 5):
            c = ws_data.cell(row=total_row_idx, column=col_idx)
            c.border = border_total
            c.fill = fill_total

        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(max_len + 4, 12)

        ws_dash.cell(row=2, column=2, value="Аналитика расходов трекера").font = font_title

        ws_dash.cell(row=4, column=2, value="Категория").font = font_header
        ws_dash.cell(row=4, column=2).fill = fill_header
        ws_dash.cell(row=4, column=2).alignment = align_center

        ws_dash.cell(row=4, column=3, value="Сумма").font = font_header
        ws_dash.cell(row=4, column=3).fill = fill_header
        ws_dash.cell(row=4, column=3).alignment = align_center

        unique_categories = sorted(list(set(exp.category for exp in self.expenses)))

        for idx, cat in enumerate(unique_categories, 5):
            c_cat = ws_dash.cell(row=idx, column=2, value=cat)
            c_cat.font = font_regular
            c_cat.border = border_data

            c_sum = ws_dash.cell(row=idx, column=3,
                                 value=f"=SUMIF('Все расходы'!C2:C{total_row_idx - 1}, B{idx}, 'Все расходы'!D2:D{total_row_idx - 1})")
            c_sum.font = font_regular
            c_sum.border = border_data
            c_sum.alignment = align_right
            c_sum.number_format = '#,##0" руб."'

            if idx % 2 == 0:
                c_cat.fill = fill_zebra
                c_sum.fill = fill_zebra

            dash_total_idx = idx

        dash_total_idx += 1
        ws_dash.cell(row=dash_total_idx, column=2, value="Всего").font = font_bold
        ws_dash.cell(row=dash_total_idx, column=2).border = border_total
        ws_dash.cell(row=dash_total_idx, column=2).fill = fill_total

        dash_total_val = ws_dash.cell(row=dash_total_idx, column=3, value=f"=SUM(C5:C{dash_total_idx - 1})")
        dash_total_val.font = font_bold
        dash_total_val.border = border_total
        dash_total_val.fill = fill_total
        dash_total_val.alignment = align_right
        dash_total_val.number_format = '#,##0" руб."'

        ws_dash.column_dimensions['B'].width = 25
        ws_dash.column_dimensions['C'].width = 18

        pie = PieChart()
        labels = Reference(ws_dash, min_col=2, min_row=5, max_row=dash_total_idx - 1)
        data = Reference(ws_dash, min_col=3, min_row=4, max_row=dash_total_idx - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Распределение расходов по категориям"
        pie.width = 16
        pie.height = 11
        ws_dash.add_chart(pie, "E4")

        wb.save(export_filename)
        return True




    def load_data(self):
        if os.path.exists(self.filename):
            with open(self.filename, "r", encoding="utf-8") as file:
                expenses = json.load(file)
                print(f"[Система] Данные успешно загружены! Найдено записей: {len(expenses)}")
                return [Expense(item["amount"], item["category"], item.get("date", "Дата не указана")) for item in expenses]
        else:
            print("[Система] База данных не найдена. Создан новый пустой список.")
            return []


    def save_data(self):
        with open(self.filename, 'w', encoding="utf-8") as file:
            ready_to_save = [asdict(item) for item in self.expenses]
            json.dump(ready_to_save, file, indent=4, ensure_ascii=False)

    def add_expense(self, amount, category):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
        item = Expense(amount, category, current_time)
        self.expenses.append(item)
        self.save_data()
        return item

    def delete_expense(self, index):
        if index <= 0:
            return False
        try:
            self.expenses.pop(index-1)
            self.save_data()
            return True
        except IndexError:
            return False

    def get_total(self):
        return sum(item.amount for item in self.expenses)

    def get_total_by_category(self, category):
        return sum(item.amount for item in self.expenses if item.category == category)

    def get_today_total(self):
        today_str = datetime.now().strftime("%Y-%m-%d")
        return sum(item.amount for item in self.expenses if item.date.startswith(today_str))


